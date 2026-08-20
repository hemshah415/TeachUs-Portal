import sys
import os
import json
import re
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def validate_email(email):
    if pd.isna(email) or not str(email).strip():
        return True # Checked in missing value rule if mandatory
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return bool(re.match(pattern, str(email).strip()))

def validate_mobile(mobile):
    if pd.isna(mobile) or not str(mobile).strip():
        return True
    clean_mob = re.sub(r'\D', '', str(mobile))
    return len(clean_mob) == 10

def safe_val(val):
    if isinstance(val, pd.Series):
        val = val.iloc[0]
    if pd.isna(val):
        return ""
    return str(val).strip()

def safe_num(val):
    if isinstance(val, pd.Series):
        val = val.iloc[0]
    if pd.isna(val):
        return None
    return val

def run_validation(file_path):
    errors = []
    students = []
    
    if not os.path.exists(file_path):
        return {
            "status": "Failed",
            "total_rows": 0,
            "passed_rows": 0,
            "error_count": 1,
            "errors": [{"row": 0, "column": "File", "error": "File does not exist", "severity": "ERROR"}],
            "students": []
        }
        
    try:
        xls = pd.ExcelFile(file_path)
    except Exception as e:
        return {
            "status": "Failed",
            "total_rows": 0,
            "passed_rows": 0,
            "error_count": 1,
            "errors": [{"row": 0, "column": "File", "error": f"Invalid Excel file format: {str(e)}", "severity": "ERROR"}],
            "students": []
        }
        
    # 1. Sheet Validation
    sheet_name = "Student_Data"
    if sheet_name not in xls.sheet_names:
        sheet_name = xls.sheet_names[0] # Fallback to first sheet
        
    # 2. Header & Row Identification
    df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    header_idx = None
    target_keywords = ["roll_number", "roll no", "student_name", "student name", "name", "email", "mobile_number", "mobile"]
    
    for idx in range(min(15, len(df_raw))):
        row_str = " ".join([str(val).lower() for val in df_raw.iloc[idx].values])
        matches = sum(1 for kw in target_keywords if kw in row_str)
        if matches >= 2:
            header_idx = idx
            break
            
    if header_idx is None:
        header_idx = 0
        
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_idx)
    
    # Standardize Column Names Map
    column_mapping = {}
    for col in df.columns:
        col_clean = str(col).strip().lower().replace(" ", "_").replace(".", "")
        if "roll" in col_clean:
            column_mapping[col] = "Roll_Number"
        elif "student_name" in col_clean or "name_of" in col_clean or col_clean == "name":
            column_mapping[col] = "Student_Name"
        elif "branch" in col_clean or "course" in col_clean or "section" in col_clean or "program" in col_clean:
            column_mapping[col] = "Branch"
        elif "sem" in col_clean:
            column_mapping[col] = "Semester"
        elif "year" in col_clean and "academic" not in col_clean:
            column_mapping[col] = "Year"
        elif "gender" in col_clean:
            column_mapping[col] = "Gender"
        elif "dob" in col_clean or "birth" in col_clean:
            column_mapping[col] = "DOB"
        elif "email" in col_clean:
            column_mapping[col] = "Email"
        elif "mobile" in col_clean or "phone" in col_clean or "contact" in col_clean:
            column_mapping[col] = "Mobile_Number"
        elif "cgpa" in col_clean:
            column_mapping[col] = "CGPA"
        elif "percent" in col_clean:
            column_mapping[col] = "Percentage"
        elif "enroll" in col_clean or "id" in col_clean or "sr" in col_clean:
            column_mapping[col] = "Enrollment_Number"
            
    df = df.rename(columns=column_mapping)
    
    # 3. Column Presence Check
    required_cols = ["Roll_Number", "Student_Name", "Branch"]
    missing_req_cols = [c for c in required_cols if c not in df.columns]
    if missing_req_cols:
        for mc in missing_req_cols:
            errors.append({
                "row": header_idx + 1,
                "column": mc,
                "error": f"Required column '{mc}' is missing in uploaded Excel sheet.",
                "severity": "ERROR"
            })
            
    seen_roll_numbers = {}
    seen_enrollments = {}
    
    total_rows = len(df)
    
    for row_num, (df_idx, row) in enumerate(df.iterrows(), start=header_idx + 2):
        row_has_error = False
        
        if row.dropna().empty:
            continue
            
        roll = safe_val(row.get("Roll_Number"))
        name = safe_val(row.get("Student_Name"))
        branch = safe_val(row.get("Branch"))
        sem = safe_num(row.get("Semester"))
        year = safe_num(row.get("Year"))
        gender = safe_val(row.get("Gender"))
        dob = safe_val(row.get("DOB"))
        email = safe_val(row.get("Email"))
        mobile = safe_val(row.get("Mobile_Number"))
        cgpa = safe_num(row.get("CGPA"))
        percentage = safe_num(row.get("Percentage"))
        enrollment = safe_val(row.get("Enrollment_Number"))

        # Mandatory Checks
        if not roll or roll.lower() in ["nan", "none"]:
            errors.append({"row": row_num, "column": "Roll_Number", "error": "Roll Number cannot be empty", "severity": "ERROR"})
            row_has_error = True
        elif roll in seen_roll_numbers:
            errors.append({"row": row_num, "column": "Roll_Number", "error": f"Duplicate Roll Number '{roll}' (First seen in row {seen_roll_numbers[roll]})", "severity": "ERROR"})
            row_has_error = True
        else:
            seen_roll_numbers[roll] = row_num

        if not name or name.lower() in ["nan", "none"]:
            errors.append({"row": row_num, "column": "Student_Name", "error": "Student Name cannot be empty", "severity": "ERROR"})
            row_has_error = True

        if not branch or branch.lower() in ["nan", "none"]:
            errors.append({"row": row_num, "column": "Branch", "error": "Branch / Course cannot be empty", "severity": "ERROR"})
            row_has_error = True

        # Semester bounds
        if pd.notna(sem):
            try:
                sem_val = int(float(sem))
                if sem_val < 1 or sem_val > 8:
                    errors.append({"row": row_num, "column": "Semester", "error": f"Invalid Semester '{sem}'. Must be between 1 and 8", "severity": "ERROR"})
                    row_has_error = True
            except ValueError:
                errors.append({"row": row_num, "column": "Semester", "error": f"Invalid Semester value '{sem}'", "severity": "ERROR"})
                row_has_error = True

        # Year bounds
        if pd.notna(year):
            try:
                yr_val = int(float(year))
                if yr_val < 1 or yr_val > 4:
                    errors.append({"row": row_num, "column": "Year", "error": f"Invalid Year '{year}'. Must be between 1 and 4", "severity": "ERROR"})
                    row_has_error = True
            except ValueError:
                errors.append({"row": row_num, "column": "Year", "error": f"Invalid Year value '{year}'", "severity": "ERROR"})
                row_has_error = True

        # Email check
        if email and not validate_email(email):
            errors.append({"row": row_num, "column": "Email", "error": f"Invalid Email address format '{email}'", "severity": "ERROR"})
            row_has_error = True

        # Mobile check
        if mobile and not validate_mobile(mobile):
            errors.append({"row": row_num, "column": "Mobile_Number", "error": f"Invalid Mobile Number '{mobile}'. Must be 10 digits", "severity": "ERROR"})
            row_has_error = True

        # CGPA bounds
        if pd.notna(cgpa):
            try:
                cgpa_val = float(cgpa)
                if cgpa_val < 0.0 or cgpa_val > 10.0:
                    errors.append({"row": row_num, "column": "CGPA", "error": f"Invalid CGPA '{cgpa}'. Must be between 0.0 and 10.0", "severity": "ERROR"})
                    row_has_error = True
            except ValueError:
                errors.append({"row": row_num, "column": "CGPA", "error": f"Invalid CGPA numeric format '{cgpa}'", "severity": "ERROR"})
                row_has_error = True

        # Percentage bounds
        if pd.notna(percentage):
            try:
                pct_val = float(percentage)
                if pct_val < 0.0 or pct_val > 100.0:
                    errors.append({"row": row_num, "column": "Percentage", "error": f"Invalid Percentage '{percentage}'. Must be between 0 and 100", "severity": "ERROR"})
                    row_has_error = True
            except ValueError:
                errors.append({"row": row_num, "column": "Percentage", "error": f"Invalid Percentage format '{percentage}'", "severity": "ERROR"})
                row_has_error = True

        # Enrollment uniqueness
        if enrollment:
            if enrollment in seen_enrollments:
                errors.append({"row": row_num, "column": "Enrollment_Number", "error": f"Duplicate Enrollment Number '{enrollment}' (First seen in row {seen_enrollments[enrollment]})", "severity": "ERROR"})
                row_has_error = True
            else:
                seen_enrollments[enrollment] = row_num

        if not row_has_error:
            students.append({
                "roll_number": roll,
                "student_name": name,
                "branch": branch,
                "semester": int(float(sem)) if pd.notna(sem) and str(sem).replace('.','',1).isdigit() else 1,
                "year": int(float(year)) if pd.notna(year) and str(year).replace('.','',1).isdigit() else 1,
                "gender": gender,
                "dob": dob,
                "email": email,
                "mobile_number": re.sub(r'\D', '', mobile),
                "cgpa": float(cgpa) if pd.notna(cgpa) and str(cgpa).replace('.','',1).isdigit() else None,
                "percentage": float(percentage) if pd.notna(percentage) and str(percentage).replace('.','',1).isdigit() else None,
                "enrollment_number": enrollment
            })

    passed_rows = len(students)
    status = "Passed" if len(errors) == 0 else "Failed"

    return {
        "status": status,
        "total_rows": total_rows,
        "passed_rows": passed_rows,
        "error_count": len(errors),
        "errors": errors,
        "students": students
    }

def generate_error_report(file_path, errors, output_error_excel):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation_Errors"
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = ["Row", "Column", "Error Message", "Severity"]
    ws.row_dimensions[1].height = 25
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    for r_idx, err in enumerate(errors, start=2):
        ws.row_dimensions[r_idx].height = 22
        c1 = ws.cell(row=r_idx, column=1, value=err.get("row"))
        c2 = ws.cell(row=r_idx, column=2, value=err.get("column"))
        c3 = ws.cell(row=r_idx, column=3, value=err.get("error"))
        c4 = ws.cell(row=r_idx, column=4, value=err.get("severity", "ERROR"))
        
        for c in [c1, c2, c3, c4]:
            c.border = thin_border
            c.alignment = Alignment(vertical="center")
            
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 15
    
    os.makedirs(os.path.dirname(output_error_excel), exist_ok=True)
    wb.save(output_error_excel)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python validate.py <input_excel_path> [output_json_path] [output_error_excel]")
        sys.exit(1)
        
    input_file = sys.argv[1]
    output_json = sys.argv[2] if len(sys.argv) > 2 else None
    output_error_excel = sys.argv[3] if len(sys.argv) > 3 else None
    
    result = run_validation(input_file)
    
    if output_error_excel and result["error_count"] > 0:
        generate_error_report(input_file, result["errors"], output_error_excel)
        result["report_path"] = output_error_excel
        
    if output_json:
        os.makedirs(os.path.dirname(output_json), exist_ok=True)
        with open(output_json, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2)
    else:
        print(json.dumps(result, indent=2))
