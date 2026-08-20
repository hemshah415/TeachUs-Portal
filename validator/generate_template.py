import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_official_template(output_path):
    wb = openpyxl.Workbook()
    
    # 1. Instructions Sheet
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr.views.sheetView[0].showGridLines = True
    
    # Title
    ws_instr.merge_cells("A1:G1")
    title_cell = ws_instr["A1"]
    title_cell.value = "COLLEGE ACADEMIC DATA SUBMISSION TEMPLATE - INSTRUCTIONS"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_instr.row_dimensions[1].height = 35
    
    instructions = [
        ("Rule #", "Requirement Description", "Example / Format"),
        ("1", "Do NOT alter the sheet name 'Student_Data' or header row in the data sheet.", "Keep sheet name exactly as 'Student_Data'"),
        ("2", "Roll_Number must be unique for each student in the file.", "2026-BMS-001"),
        ("3", "Mandatory fields: Roll_Number, Student_Name, Branch, Semester, Year, Email, Mobile_Number.", "All must be filled"),
        ("4", "Semester must be a number between 1 and 8.", "1, 2, 3, etc."),
        ("5", "Year must be a number between 1 and 4.", "1, 2, 3, 4"),
        ("6", "Mobile_Number must be a 10-digit number.", "9820012345"),
        ("7", "Email must be a valid email address format.", "student@college.edu.in"),
        ("8", "CGPA must be between 0.00 and 10.00. Percentage must be between 0.00 and 100.00.", "8.75 / 85.50"),
        ("9", "Date of Birth (DOB) format should be YYYY-MM-DD.", "2005-04-15"),
    ]
    
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    for r_idx, row in enumerate(instructions, start=3):
        ws_instr.row_dimensions[r_idx].height = 24
        for c_idx, val in enumerate(row, start=1):
            cell = ws_instr.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if r_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
                
    ws_instr.column_dimensions['A'].width = 10
    ws_instr.column_dimensions['B'].width = 65
    ws_instr.column_dimensions['C'].width = 35
    
    # 2. Student Data Sheet
    ws_data = wb.create_sheet(title="Student_Data")
    ws_data.views.sheetView[0].showGridLines = True
    
    headers = [
        "Roll_Number", "Student_Name", "Branch", "Semester", "Year",
        "Gender", "DOB", "Email", "Mobile_Number", "CGPA", "Percentage", "Enrollment_Number"
    ]
    
    ws_data.row_dimensions[1].height = 30
    for c_idx, header in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=c_idx, value=header)
        cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    sample_rows = [
        ["NKC-2026-001", "Aarav Sharma", "BMS", 1, 1, "Male", "2006-05-12", "aarav.sharma@example.com", "9820123456", 8.90, 84.50, "ENR2026001"],
        ["NKC-2026-002", "Ananya Patel", "BSC IT", 1, 1, "Female", "2006-08-20", "ananya.patel@example.com", "9820234567", 9.20, 88.00, "ENR2026002"],
        ["NKC-2026-003", "Rohan Mehta", "FYBCOM", 1, 1, "Male", "2005-11-03", "rohan.mehta@example.com", "9820345678", 7.80, 75.20, "ENR2026003"]
    ]
    
    for r_idx, row_data in enumerate(sample_rows, start=2):
        ws_data.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"[OK] Template generated successfully at: {output_path}")

if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\Admin\Desktop\Collegedatamanagementsystem\backend\uploads\Official_Academic_Data_Template.xlsx"
    generate_official_template(out)
